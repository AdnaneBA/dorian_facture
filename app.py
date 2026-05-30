import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import io
import copy

# ─────────────────────────────────────────────
# COLONNES CLÉS pour la comparaison (modifiable)
# ─────────────────────────────────────────────
CLE_COLONNES = ["N° de pièce", "Date fact Frs", "Montant Devise"]

# ─────────────────────────────────────────────
# CONFIG PAGE
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Validation Factures",
    page_icon="🧾",
    layout="centered",
)

# ─────────────────────────────────────────────
# CSS CUSTOM
# ─────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

    html, body, [class*="css"] {
        font-family: 'DM Sans', sans-serif;
    }

    .stApp {
        background: #f7f6f3;
    }

    .header-block {
        background: #1a1a2e;
        border-radius: 16px;
        padding: 2rem 2.5rem;
        margin-bottom: 2rem;
        color: white;
    }

    .header-block h1 {
        font-size: 1.8rem;
        font-weight: 600;
        margin: 0 0 0.3rem 0;
        letter-spacing: -0.02em;
    }

    .header-block p {
        font-size: 0.95rem;
        color: #a0a0c0;
        margin: 0;
    }

    .upload-section {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1rem;
        border: 1px solid #e8e8e4;
    }

    .upload-label {
        font-size: 0.78rem;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #888;
        margin-bottom: 0.5rem;
    }

    .tag-green {
        display: inline-block;
        background: #d4f5e4;
        color: #1a7a4a;
        border-radius: 6px;
        padding: 2px 10px;
        font-size: 0.78rem;
        font-weight: 600;
    }

    .tag-blue {
        display: inline-block;
        background: #dbeafe;
        color: #1d4ed8;
        border-radius: 6px;
        padding: 2px 10px;
        font-size: 0.78rem;
        font-weight: 600;
    }

    .stat-card {
        background: white;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        border: 1px solid #e8e8e4;
        text-align: center;
    }

    .stat-number {
        font-size: 2.2rem;
        font-weight: 600;
        font-family: 'DM Mono', monospace;
        line-height: 1;
    }

    .stat-label {
        font-size: 0.8rem;
        color: #888;
        margin-top: 0.3rem;
    }

    .alert-ok {
        background: #d4f5e4;
        border-left: 4px solid #1a7a4a;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        color: #1a7a4a;
        font-weight: 500;
    }

    .alert-warn {
        background: #fff3cd;
        border-left: 4px solid #e6a817;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        color: #7a4f00;
        font-weight: 500;
    }

    .cle-block {
        background: #f0f0f8;
        border-radius: 8px;
        padding: 0.8rem 1.2rem;
        font-family: 'DM Mono', monospace;
        font-size: 0.85rem;
        color: #444;
        margin-top: 0.5rem;
    }

    div[data-testid="stFileUploader"] {
        border-radius: 10px;
    }

    .stDownloadButton button {
        background: #1a1a2e !important;
        color: white !important;
        border-radius: 10px !important;
        padding: 0.6rem 1.5rem !important;
        font-family: 'DM Sans', sans-serif !important;
        font-weight: 500 !important;
        font-size: 0.95rem !important;
        border: none !important;
        width: 100%;
        transition: background 0.2s;
    }

    .stDownloadButton button:hover {
        background: #2d2d4e !important;
    }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="header-block">
    <h1>🧾 Validation Factures</h1>
    <p>Comparez deux fichiers Excel et surlignez automatiquement les factures déjà validées du mois précédent.</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# COLONNES CLÉS utilisées
# ─────────────────────────────────────────────
with st.expander("⚙️ Colonnes clés utilisées pour la comparaison", expanded=False):
    st.markdown("Les lignes sont considérées comme identiques si ces colonnes correspondent :")
    cles_str = "  +  ".join(CLE_COLONNES)
    st.markdown(f'<div class="cle-block">🔑 {cles_str}</div>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# UPLOAD
# ─────────────────────────────────────────────
col1, col2 = st.columns(2)

with col1:
    st.markdown('<div class="upload-label">📁 Fichier M-1 (mois précédent)</div>', unsafe_allow_html=True)
    file_m1 = st.file_uploader("", type=["xlsx", "xls"], key="m1", label_visibility="collapsed")
    if file_m1:
        st.markdown('<span class="tag-green">✓ Chargé</span>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="upload-label">📁 Fichier M (mois actuel)</div>', unsafe_allow_html=True)
    file_m = st.file_uploader("", type=["xlsx", "xls"], key="m", label_visibility="collapsed")
    if file_m:
        st.markdown('<span class="tag-blue">✓ Chargé</span>', unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# TRAITEMENT
# ─────────────────────────────────────────────
def normaliser(valeur):
    """Normalise une valeur pour la comparaison (gère NaN, espaces, casse)."""
    if pd.isna(valeur):
        return ""
    return str(valeur).strip().lower()

def construire_cle(row, colonnes):
    return tuple(normaliser(row[c]) for c in colonnes if c in row.index)

def traiter_fichiers(file_m1, file_m):
    # Lecture des données
    df_m1 = pd.read_excel(file_m1, dtype=str)
    df_m = pd.read_excel(file_m, dtype=str)

    # Vérification que les colonnes clés existent
    cols_manquantes_m1 = [c for c in CLE_COLONNES if c not in df_m1.columns]
    cols_manquantes_m = [c for c in CLE_COLONNES if c not in df_m.columns]

    if cols_manquantes_m1 or cols_manquantes_m:
        return None, None, None, cols_manquantes_m1, cols_manquantes_m

    # Construire l'ensemble des clés du fichier M-1
    cles_m1 = set()
    for _, row in df_m1.iterrows():
        cles_m1.add(construire_cle(row, CLE_COLONNES))

    # Identifier les lignes de M qui correspondent à M-1
    lignes_a_surligner = []
    for idx, row in df_m.iterrows():
        cle = construire_cle(row, CLE_COLONNES)
        if cle in cles_m1:
            lignes_a_surligner.append(idx)

    # Appliquer le surlignage vert dans le fichier Excel
    file_m.seek(0)
    wb = load_workbook(file_m)
    ws = wb.active

    # Couleur vert doux
    fill_vert = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # La ligne 1 est l'en-tête → les données commencent à la ligne 2
    # df.index commence à 0 → ligne Excel = index + 2
    for idx in lignes_a_surligner:
        ligne_excel = idx + 2
        for col in range(1, ws.max_column + 1):
            ws.cell(row=ligne_excel, column=col).fill = fill_vert

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, len(lignes_a_surligner), len(df_m1), [], []

# ─────────────────────────────────────────────
# BOUTON LANCER
# ─────────────────────────────────────────────
if file_m1 and file_m:
    if st.button("▶ Lancer la comparaison", use_container_width=True):
        with st.spinner("Analyse en cours..."):
            resultat, nb_surligne, nb_m1, err_m1, err_m = traiter_fichiers(file_m1, file_m)

        if err_m1 or err_m:
            st.error(f"❌ Colonnes introuvables dans M-1 : {err_m1}" if err_m1 else "")
            st.error(f"❌ Colonnes introuvables dans M : {err_m}" if err_m else "")
        else:
            st.markdown("<br>", unsafe_allow_html=True)

            # Stats
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number" style="color:#1d4ed8">{nb_m1}</div>
                    <div class="stat-label">Lignes dans M-1</div>
                </div>""", unsafe_allow_html=True)
            with c2:
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number" style="color:#1a7a4a">{nb_surligne}</div>
                    <div class="stat-label">Lignes surlignées dans M</div>
                </div>""", unsafe_allow_html=True)
            with c3:
                delta = nb_m1 - nb_surligne
                couleur = "#1a7a4a" if delta == 0 else "#e6a817"
                st.markdown(f"""
                <div class="stat-card">
                    <div class="stat-number" style="color:{couleur}">{delta}</div>
                    <div class="stat-label">Lignes non retrouvées</div>
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Alerte cohérence
            if nb_surligne == nb_m1:
                st.markdown("""
                <div class="alert-ok">
                    ✅ Parfait — Toutes les lignes de M-1 ont été retrouvées et surlignées dans M.
                </div>""", unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="alert-warn">
                    ⚠️ Attention — {nb_m1 - nb_surligne} ligne(s) de M-1 n'ont pas été retrouvées dans M.
                    Vérifiez que les colonnes clés sont identiques entre les deux fichiers.
                </div>""", unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # Téléchargement
            import os
            base = os.path.splitext(file_m.name)[0]
            nom_fichier = f"{base}_validé.xlsx"
            st.download_button(
                label="⬇️ Télécharger le fichier M avec surlignage",
                data=resultat,
                file_name=nom_fichier,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
else:
    st.info("👆 Chargez les deux fichiers Excel pour commencer.")
